import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook
import re
from datetime import datetime, timedelta
import os
from icecream import ic
from bs4 import BeautifulSoup

def browse_files():
    # Get the path to the Downloads folder
    downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
    report_folder = r'G:\Shared drives\O&M\NCC Automations\In Progress'
    # Open file dialog to select the first file
    wo_data = filedialog.askopenfilename(
        title="Select the Emaint Report Download Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")],
        initialdir=downloads_folder
    )
    if not wo_data:
        return
    
    # Open file dialog to select the second file
    destination = filedialog.askopenfilename(title="Select the Destination Excel file", filetypes=[("Excel files", "*.xlsx *.xls")], initialdir=report_folder)
    if not destination:
        return
    
    # Process the files
    process_files(wo_data, destination)

def calculate_hours_between(start_datetime, end_datetime):
    start_hour = 7
    end_hour = 20
    total_minutes = 0

    current_datetime = start_datetime

    while current_datetime < end_datetime:
        if start_hour <= current_datetime.hour < end_hour:
            total_minutes += 1
        current_datetime += timedelta(minutes=1)

    total_hours = round(total_minutes / 60, 2)
    return total_hours

def append_data_to_excel(file_path, marshallData, ogburnData, jeffersonData, tedderData, hicksonData, bishopvilleIIData, vanburenData, thunderheadData):
    # Load the workbook and select the active sheet
    workbook = load_workbook(file_path)
    vanburenSheet = workbook['Van Buren']
    thunderheadSheet = workbook['Thunderhead']
    tedderSheet = workbook['Tedder']
    marshallSheet = workbook['Marshall']
    ogburnSheet = workbook['Ogburn']
    hicksonSheet = workbook['Hickson']
    jeffersonSheet = workbook['Jefferson']
    bishopvilleIISheet = workbook['Bishopville']

    # Determine the starting row
    start_row = 9
    end_row = 100
    # Function to clear contents from a specific range in a sheet
    def clear_sheet_range(sheet, start_row, end_row):
        for row in range(start_row, end_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
    # Clear contents from each sheet
    clear_sheet_range(vanburenSheet, start_row, end_row)
    clear_sheet_range(thunderheadSheet, start_row, end_row)
    clear_sheet_range(tedderSheet, start_row, end_row)
    clear_sheet_range(marshallSheet, start_row, end_row)
    clear_sheet_range(ogburnSheet, start_row, end_row)
    clear_sheet_range(hicksonSheet, start_row, end_row)
    clear_sheet_range(jeffersonSheet, start_row, end_row)
    clear_sheet_range(bishopvilleIISheet, start_row, end_row)


    # Function to append data to a specific sheet
    def append_to_sheet(sheet, data):
        for i, row in enumerate(data, start=start_row):
            for j, value in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=value)

    # Append data to each sheet
    append_to_sheet(vanburenSheet, vanburenData)
    append_to_sheet(thunderheadSheet, thunderheadData)
    append_to_sheet(tedderSheet, tedderData)
    append_to_sheet(marshallSheet, marshallData)
    append_to_sheet(ogburnSheet, ogburnData)
    append_to_sheet(hicksonSheet, hicksonData)
    append_to_sheet(jeffersonSheet, jeffersonData)
    append_to_sheet(bishopvilleIISheet, bishopvilleIIData)

    # Save the workbook
    workbook.save(file_path)
    #Open
    os.startfile(file_path)

def process_files(wo_data, destination):
    # Load the first workbook
    wb1 = load_workbook(wo_data)
    ws1 = wb1.active


    vanburenData = []
    thunderheadData = []
    tedderData = []
    marshallData = []
    ogburnData = []
    hicksonData = []
    jeffersonData = []
    bishopvilleIIData = []

    # Example: Copy data from the first workbook to the second with some formatting
    for index, row in enumerate(ws1.iter_rows(values_only=True)):
        if index == 0:
            continue
        #Reset Variables
        wo = None
        kind = None
        startdate = None
        starttime = None
        wo_date = None
        endtime = None
        cellG = None
        duration_in_hours = None
        prod_hours = None
        
        wo = row[0] #Cell 1 of Destination Sheet Row
        site = row[2]
        wo_date = row[4]
        description = BeautifulSoup(row[8], "html.parser").get_text()
        complete = row[9]
        #This will get overwritten and if not then thats ideal 
        startdate = datetime.strftime(wo_date, '%Y-%m-%d')
        
        #Cell 2.  Search for the words "Stow" and "Curtailment" in the description

        invQ = re.search(r'(inverter|transformer|CB|Combiner Box|medium voltage)', description, re.IGNORECASE)
        if invQ:
            kind = 'Inverter'
        else:
            stowQ = re.search(r'(Stow|Curtail)', description, re.IGNORECASE)
            if stowQ:
                step2 = stowQ.group(0)
                if step2 == 'stow':
                    kind = 'Stow'
                elif step2 == 'curtail':
                    kind = 'Curtailment'
            else:
                kind = "Outage"
        #Concluded

        #Start Date & End Date
        startdate_match = re.search(r'Start Date: (\d{1,2}[-/]\d{1,2}(?:[-/]\d{2,4})?)', description)
        starttime_match = re.search(r'Start Time: (\d{1,2}[:;]?\d{1,2}|\d{3,4})', description)
        enddate_match = re.search(r'End Date: (\d{1,2}[-/]\d{1,2}(?:[-/]\d{2,4})?)', description)
        endtime_match = re.search(r'End Time: (\d{1,2}[:;]?\d{1,2}|\d{3,4})', description)
        # Function to add current year if missing
        def add_current_year(date_str):
            # Check if the year is missing
            if re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{4}', date_str):
                return date_str  # Year is present
            elif re.match(r'\d{1,2}[-/]\d{1,2}[-/]\d{2}', date_str):
                parts = re.split(r'[-/]', date_str)
                month, day, year = parts[0], parts[1], parts[2]
                fyear = int(year) + 2000
                return f"{month}/{day}/{fyear}"
            else:
                current_year = datetime.now().year
                return f"{date_str}/{current_year}"
        def format_time(time_str):
            # If the time is already in HH:MM format, return it as is
            if ':' in time_str:
                return time_str
            # If the time is a group of 3 or 4 digits, format it to HH:MM
            elif len(time_str) == 3:
                return f"0{time_str[0]}:{time_str[1:]}"
            elif len(time_str) == 4:
                return f"{time_str[:2]}:{time_str[2:]}"
            else:
                raise ValueError("Invalid time format")

        # Extract and format the start time
        if starttime_match:
            print(starttime_match.group(1), description)
            starttime = format_time(starttime_match.group(1))
            print("Formatted Start Time:", starttime)

        # Extract and format the end time
        if endtime_match:
            endtime = format_time(endtime_match.group(1))
            print("Formatted End Time:", endtime)

        # Extract the end date
        if enddate_match:
            enddate = add_current_year(enddate_match.group(1))
        # Extract the end date
        if startdate_match:
            startdate = add_current_year(startdate_match.group(1))

        start_datetime = datetime.strptime(f"{startdate} {starttime}", '%m/%d/%Y %H:%M')
        end_datetime = datetime.strptime(f"{enddate} {endtime}", '%m/%d/%Y %H:%M')
        
        # Calculate the duration
        if start_datetime and end_datetime:
            duration = end_datetime - start_datetime
            duration_in_hours = round(duration.total_seconds() / 3600, 2)
        else:
            duration = None

        #Production Hours
        if start_datetime and end_datetime:
            prod_hours = calculate_hours_between(start_datetime, end_datetime)



        #Cell 5
        remote_search = re.search(r'(remote|NCC Closed)', complete, re.IGNORECASE)
        if remote_search:
            cellG = "Yes"
        else:
            remote_search2 = re.search(r'NCC Closed', description, re.IGNORECASE)
            if remote_search2:
                cellG = "Yes"
            else:
                cellG = ""


        data = (wo, kind, startdate, starttime, wo_date, endtime, cellG, duration_in_hours, prod_hours)
        if site == "Marshall":
            #ic(data)
            marshallData.append(data)
        elif site == "OGBURN":
            ogburnData.append(data)
        elif site == "JEFFERSON":
            jeffersonData.append(data)
        elif site == "Tedder":
            tedderData.append(data)
        elif site == "HICKSON":
            hicksonData.append(data)
        elif site == "BISHOPVILLE":
            bishopvilleIIData.append(data)
        elif site == "Van Buren":
            vanburenData.append(data)
        elif site == "Thunderhead":
            thunderheadData.append(data)


    #Here
    append_data_to_excel(destination, marshallData, ogburnData, jeffersonData, tedderData, hicksonData, bishopvilleIIData, vanburenData, thunderheadData)
    root.destroy()

        


# Set up the tkinter window
root = tk.Tk()
root.title("Harrison St. Outage Processing")

# Create a button to browse files
browse_button = tk.Button(root, text="Browse Files", command=browse_files, height=5, width=50)
browse_button.pack()

# Run the tkinter main loop
root.mainloop()